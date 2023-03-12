# -*- coding: utf-8 -*-
import pandas as pd
import glob
import openpyxl
import csv
import os
import shutil
import os.path
import uuid
from openpyxl import load_workbook
from openpyxl.chart import LineChart, PieChart, BarChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule


NUM_MATCHES = 99
SUM_COMMUNITY = 100
SUM_NO_COMMUNITY = 101
SUM_DOCK_AUTON = 102
SUM_ENGAGE_AUTON = 103
SUM_NEITHER_AUTON = 104
SUM_DOCK_TELEOP = 105
SUM_ENGAGE_TELEOP = 106
SUM_NEITHER_TELEOP = 107
SUM_CONE_AUTON = 108
SUM_CUBE_AUTON = 109
SUM_CONE_TELEOP = 110
SUM_CUBE_TELEOP = 111

FILE_NAME = "./scouting_app_data_interface.xlsx"
tablet_data = glob.glob("./data_dump/*")
book = openpyxl.load_workbook(FILE_NAME)

def create_new_sheet(list_of_elem, sheet):
    #makes the data into a dataframe to place inside of sheet
    df_list = pd.DataFrame([list_of_elem])
    
    #initializes sheet to write in so we can do the thing
    writer = pd.ExcelWriter(FILE_NAME, engine = 'openpyxl')
    writer.book = book
    
    #makes new worksheet within workbook with dataframe as data
    try:
        df_list.to_excel(writer, sheet_name = sheet, index = False, header=["Name", "Match", "Team", "Tablet", "community", "gamepiece-type", "gamepiece-type3", "auton-upper-node-count",     "auton-middle-node-count", "auton-lower-node-count", "dock", "engage", "neither", "telop-upper-node-count", "telop-middle-node-count", "telop-lower-node-count", "gamepiece-type4", "gamepiece-type5", "foul-count", "yes defense", "no defense", "defense", "dock2", "engage2", "neither2", "yesbreak", "nobreak", "comments", "empty"])
    except:
        print("\n ERROR doing " + sheet)
        return
        
    writer.save()
    
    
    writer = pd.ExcelWriter(FILE_NAME, engine = 'openpyxl')
    writer.book = book


    sheets = {ws.title: ws for ws in book.worksheets}
    ws = sheets[sheet]
    target = book.copy_worksheet(sheets['Summary'])
    target.title = sheet + ' Summary'
    target.cell(2, 1).value = sheet
    
   
    rule = ColorScaleRule(start_type='percentile', start_value=10, start_color='FFAA0000',
                       mid_type='percentile', mid_value=50, mid_color='FFAAAA00',
                       end_type='percentile', end_value=90, end_color='FF00AA00')
    target.conditional_formatting.add('C28:C31', rule)
    target.conditional_formatting.add('E28:E30', rule)
    target.conditional_formatting.add('G28:G30', rule)
    target.conditional_formatting.add('H28:H28', rule)
    target.conditional_formatting.add('H31:H31', rule)
    
    
    chart = LineChart()
    chart.title = "Auton " + sheet
    chart.y_axis.title = 'Goals'
    chart.x_axis.title = 'Match Number'
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=8, min_row=1, max_col=8, max_row=21)
    chart.add_data(values, titles_from_data=True)
    values = Reference(ws, min_col=9, min_row=1, max_col=9, max_row=21)
    chart.add_data(values, titles_from_data=True)
    values = Reference(ws, min_col=10, min_row=1, max_col=10, max_row=21)
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "A22")
    
    chart = LineChart()
    chart.title = "Teleop " + sheet
    chart.y_axis.title = 'Goals'
    chart.x_axis.title = 'Match Number'
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=14, min_row=1, max_col=14, max_row=21)
    chart.add_data(values, titles_from_data=True)
    values = Reference(ws, min_col=15, min_row=1, max_col=15, max_row=21)
    chart.add_data(values, titles_from_data=True)
    values = Reference(ws, min_col=16, min_row=1, max_col=16, max_row=21)
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "J22")
    
    ws.cell(1, NUM_MATCHES).value = "Matches"
    ws.cell(2, NUM_MATCHES).value = "=COUNT($E$2:$E$20)"
    ws.cell(1, SUM_COMMUNITY).value = "Community"
    ws.cell(2, SUM_COMMUNITY).value = "=SUM($E$2:$E$20)"
    ws.cell(1, SUM_NO_COMMUNITY).value = "No Community"
    ws.cell(2, SUM_NO_COMMUNITY).value = "=$CU$2-$CV$2"
    ws.cell(1, SUM_DOCK_AUTON).value = "Dock Auton"
    ws.cell(2, SUM_DOCK_AUTON).value = "=SUM($K$2:$K$20)"
    ws.cell(1, SUM_ENGAGE_AUTON).value = "Engage Auton"
    ws.cell(2, SUM_ENGAGE_AUTON).value = "=SUM($L$2:$L$20)"
    ws.cell(1, SUM_NEITHER_AUTON).value = "Neither Auton"
    ws.cell(2, SUM_NEITHER_AUTON).value = "=SUM($M$2:$M$20)"
    ws.cell(1, SUM_DOCK_TELEOP).value = "Dock Teleop"
    ws.cell(2, SUM_DOCK_TELEOP).value = "=SUM($W$2:$W$20)"
    ws.cell(1, SUM_ENGAGE_TELEOP).value = "Engage Teleop"
    ws.cell(2, SUM_ENGAGE_TELEOP).value = "=SUM($X$2:$X$20)"
    ws.cell(1, SUM_NEITHER_TELEOP).value = "Neither Teleop"
    ws.cell(2, SUM_NEITHER_TELEOP).value = "=SUM($Y$2:$Y$20)"
    ws.cell(1, SUM_CUBE_AUTON).value = "Cube Auton"
    ws.cell(2, SUM_CUBE_AUTON).value = "=SUM($G$2:$G$20)"
    ws.cell(1, SUM_CONE_AUTON).value = "Cone Auton"
    ws.cell(2, SUM_CONE_AUTON).value = "=SUM($F$2:$F$20)"
    ws.cell(1, SUM_CUBE_TELEOP).value = "Cube Teleop"
    ws.cell(2, SUM_CUBE_TELEOP).value = "=SUM($R$2:$R$20)"
    ws.cell(1, SUM_CONE_TELEOP).value = "Cone Teleop"
    ws.cell(2, SUM_CONE_TELEOP).value = "=SUM($Q$2:$Q$20)"
    
    chart = BarChart()
    chart.title = "Community " + sheet
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=SUM_COMMUNITY, min_row=1, max_col=SUM_NO_COMMUNITY, max_row=2)
    chart.add_data(values, titles_from_data=True)
    #chart.set_categories(Reference(ws, min_col=SUM_COMMUNITY, min_row=1, max_col=SUM_NO_COMMUNITY, max_row=1))
    ws.add_chart(chart, "S22")
    
    
    chart = BarChart()
    chart.title = "Docking Auton " + sheet
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=SUM_DOCK_AUTON, min_row=1, max_col=SUM_NEITHER_AUTON, max_row=2)
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "A36")
    
    chart = BarChart()
    chart.title = "Docking Teleop " + sheet
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=SUM_DOCK_TELEOP, min_row=1, max_col=SUM_NEITHER_TELEOP, max_row=2)
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "J36")
    
    chart = BarChart()
    chart.title = "Docking Teleop " + sheet
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=SUM_DOCK_TELEOP, min_row=1, max_col=SUM_NEITHER_TELEOP, max_row=2)
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "J36")
    
    chart = BarChart()
    chart.title = "Gamepiece Auton " + sheet
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=SUM_CONE_AUTON, min_row=1, max_col=SUM_CUBE_AUTON, max_row=2)
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "A50")
    
    chart = BarChart()
    chart.title = "Gamepiece Teleop " + sheet
    chart.y_axis.scaling.max = 20
    values = Reference(ws, min_col=SUM_CONE_TELEOP, min_row=1, max_col=SUM_CUBE_TELEOP, max_row=2)
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "J50")

    writer.save()

def append_existing_sheet(list_of_elem, sheet):
    df_list = pd.DataFrame([list_of_elem])
    writer = pd.ExcelWriter(FILE_NAME, engine='openpyxl')
    writer.book = book
    sheets = {ws.title: ws for ws in book.worksheets}
    
    df_list.to_excel(writer,sheet_name=sheet, startrow=sheets[sheet].max_row, index = False,header= False)

    writer.save()

def updateMain():
    pass

def lookForValue(sheet):
    """
    Find value within a worksheet.
    """
    for row in sheet.iter_rows():
        if row[0].value == "Averages:":
            return row
        else:
            return None

# def appendAverages():
#     """
#     Appends a row of averages for each number-based value, and the majority for the Y/N values
#     """
#     for sheet in book.worksheets:  
#         if lookForValue(sheet) == None:
#             append_existing_sheet(['Averages:', '-', '-', df["AUTON_UPPER"].mean(), df["AUTON_LOWER"].mean(), df["TELEOP_UPPER"].mean(), df["TELEOP_LOWER"].mean(), df["ENDGAME_CLIMB"].value_counts().idxmax(), "-", df["TARMAC_Y/N"].value_counts().idxmax(), df["FENDER_Y/N"].value_counts().idxmax(), df["LAUNCH PAD_Y/N"].value_counts().idxmax(), df["TERMINAL_Y/N"].value_counts().idxmax(), df["MID-FIELD_Y/N"].value_counts().idxmax(), df["TDEFENSE_Y/N"].value_counts().idxmax(), df["TDEFENDED_Y/N"].value_counts().idxmax(), df["PHP_Y/N"].value_counts().idxmax(), df["PGROUND_Y/N"].value_counts().idxmax(), df["EDEFENSE"].value_counts().idxmax(), df["ESCORE"].value_counts().idxmax()], sheet)
#             pass
#         else:
#             #delete the row with averages
#             #append_existing_sheet(['Averages:', '-', '-', df["AUTON_UPPER"].mean(), df["AUTON_LOWER"].mean(), df["TELEOP_UPPER"].mean(), df["TELEOP_LOWER"].mean(), df["ENDGAME_CLIMB"].value_counts().idxmax(), "-", df["TARMAC_Y/N"].value_counts().idxmax(), df["FENDER_Y/N"].value_counts().idxmax(), df["LAUNCH PAD_Y/N"].value_counts().idxmax(), df["TERMINAL_Y/N"].value_counts().idxmax(), df["MID-FIELD_Y/N"].value_counts().idxmax(), df["TDEFENSE_Y/N"].value_counts().idxmax(), df["TDEFENDED_Y/N"].value_counts().idxmax(), df["PHP_Y/N"].value_counts().idxmax(), df["PGROUND_Y/N"].value_counts().idxmax(), df["EDEFENSE"].value_counts().idxmax(), df["ESCORE"].value_counts().idxmax()], sheet)
#             pass
            
def writeToTeamFile(tablet_data):
  """
  Creates a sheet in the excel workbook for each team scouted OR adds data to an existing file.
  """
  shutil.copy(FILE_NAME, f'./full_backups/{uuid.uuid4()}.xlsx')
  for data in tablet_data:
    print(f"Processing file {data}")
    with open(data, newline='') as csvfile:
        reader = csv.reader(csvfile)
        list_of_elem = next(reader)
    for i in range (4,27):
        list_of_elem[i] = int(list_of_elem[i])
    try:
        try:
            append_existing_sheet(list_of_elem, list_of_elem[2])
            append_existing_sheet(list_of_elem, "All Data")
        except:
            create_new_sheet(list_of_elem, list_of_elem[2])
            append_existing_sheet(list_of_elem, "All Data")
        newfile = "./backup/" + os.path.basename(data)
        os.rename(data, newfile)
    except:
        newfile = "./bad_data/" + os.path.basename(data)
        os.rename(data, newfile)

writeToTeamFile(tablet_data)
#appendAverages()

#"""This is to print out one of the dataframes to see if everything is in order"""
#print(pd.read_csv("team_data/2607.csv", names=["MATCH_NUM", "SCOUTER_NAME", "TEAM_NUM", "AUTON_UPPER", "AUTON_LOWER", "TELEOP_UPPER", "TELEOP_LOWER", "ENDGAME_CLIMB", "COMMENTS", "TARMAC_Y/N", "FENDER_Y/N", "LAUNCH PAD_Y/N", "TERMINAL_Y/N", "MID-FIELD_Y/N", "TDEFENSE_Y/N", "TDEFENDED_Y/N", "PHP_Y/N", "PGROUND_Y/N", "EDEFENSE", "ESCORE"]))
